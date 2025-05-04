# Advanced Solana Trading Bot Core Functions
# Optimized for high-performance trading operations

import os
import sys
import time
import pandas as pd
import layouts
import asyncio
import datetime
import base58
import websockets
import json
from solders.signature import Signature
import requests
from solders.pubkey import Pubkey
from solana.rpc.api import Client, Keypair
from solana.rpc.async_api import AsyncClient
from datetime import datetime
from spl.token.constants import TOKEN_PROGRAM_ID, WRAPPED_SOL_MINT
from solana.rpc import types
from solana.rpc.types import TokenAccountOpts
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
import re
import sqlite3

# Constants and Configuration
SOLANA_DECIMALS = 10**9
MAX_TOKEN_ACCOUNTS = 15000
MIN_TOKEN_ACCOUNTS = 1

class TerminalColors:
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'

# Global Constants
TRACKED_SIGNATURES = set()
SOLANA_WRAPPED_MINT = "So11111111111111111111111111111111111111112"
RAYDIUM_POOL = "675kPX9MHTjS2zt1qfr1NYHuzeLXfQM9H24wFSUt1Mp8"
RAYDIUM_V4 = "5Q544fKrFoe6tsEbD7S8EmxGTJYAKtTVhAW5Q5pge4j1"

class SolanaTrader:
    def __init__(self, wallet_address):
        self.trade_queue = asyncio.Queue()
        self.wallet_address = wallet_address
        self.active_token_accounts = 0
        self.solana_client = AsyncClient("RPC_HTTPS_URL")
        self.database_name = "trading_data.db"
        self.db_connection = sqlite3.connect(self.database_name)
        self.db_cursor = self.db_connection.cursor()
        self.initialize_database()
        self.wallet_id = self.get_wallet_identifier(wallet_address)
        
        # Trading metrics
        self.total_income = 0
        self.total_outcome = 0
        self.transaction_fees = 0
        self.sol_spent = 0
        self.sol_earned = 0
        self.buy_count = 0
        self.sell_count = 0
        self.sol_difference = 0
        self.token_difference = 0
        self.profit_percentage = 0
        self.initial_buy_time = None
        self.final_sell_time = None
        self.last_transaction_time = None
        self.trading_duration = 0
        self.current_contract = None
        self.suspicious_tokens = 0
        self.sol_balance = None
        self.solana_usd_price = self.fetch_solana_price()
        self.token_creation_time = None
        self.purchase_period = 0

    def initialize_database(self):
        self.db_cursor.execute('''
                    CREATE TABLE IF NOT EXISTS wallet_address (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        wallet_address TEXT UNIQUE
                    )
                ''')

        self.db_cursor.execute('''
                    CREATE TABLE IF NOT EXISTS token_accounts (
                        wallet_address_id INTEGER,
                        wallet_token_account TEXT,
                        block_time INTEGER,
                        FOREIGN KEY(wallet_address_id) REFERENCES wallet_address(id)
                    )
                ''')

        self.db_cursor.execute('''
                 CREATE TABLE IF NOT EXISTS pnl_info (
                     token_account TEXT PRIMARY KEY,
                     wallet_address_id INTEGER,
                     income REAL,
                     outcome REAL,
                     total_fee REAL,
                     spent_sol REAL,
                     earned_sol REAL,
                     delta_token REAL,
                     delta_sol REAL,
                     delta_percentage REAL,
                     buys INTEGER,
                     sells INTEGER,
                     last_trade TEXT,
                     time_period TEXT,
                     contract TEXT,
                     scam_tokens TEXT,
                     buy_period TEXT, 

                     FOREIGN KEY(wallet_address_id) REFERENCES wallet_address(id)
                 )
             ''')

        self.db_cursor.execute('''
                       CREATE TABLE IF NOT EXISTS winning_wallets (
                           wallet_address_id INTEGER,
                           win_rate_7 REAL,
                           balance_change_7 REAL,
                           token_accounts_7 INTEGER,
                           win_rate_14 REAL,
                           balance_change_14 REAL,
                           token_accounts_14 INTEGER,
                           win_rate_30 REAL,
                           balance_change_30 REAL,
                           token_accounts_30 INTEGER,
                           win_rate_60 REAL,
                           balance_change_60 REAL,
                           token_accounts_60 INTEGER,
                           win_rate_90 REAL,
                           balance_change_90 REAL,
                           token_accounts_90 INTEGER,
                           FOREIGN KEY(wallet_address_id) REFERENCES wallet_address(id)
                       )
                   ''')
        self.db_connection.commit()

    async def get_token_accountsCount(self, wallet_address: Pubkey):
        owner = wallet_address
        opts = types.TokenAccountOpts(program_id=Pubkey.from_string("TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA"))
        response = await self.solana_client.get_token_accounts_by_owner(owner, opts)
        return len(response.value)

    async def initialize(self):
        self.sol_balance = await self.getSOlBalance()

    async def getSOlBalance(self):
        pubkey = self.wallet_address
        response = await self.solana_client.get_balance(pubkey)
        balance = response.value / SOLANA_DECIMALS
        return balance

    def fetch_solana_price(self):
        url = "https://api.coingecko.com/api/v3/simple/price?ids=solana&vs_currencies=usd"
        response = requests.get(url)
        data = response.json()
        return data['solana']['usd'] if response.status_code == 200 else None

    def store_win_rate(self, time_period, win_rate, balance_change, token_accounts):
        check_sql = 'SELECT 1 FROM winning_wallets WHERE wallet_address_id = ?'
        self.db_cursor.execute(check_sql, (self.wallet_id,))
        row_exists = self.db_cursor.fetchone() is not None

        time_period_suffix = f"_{time_period}"

        if row_exists:
            update_sql = f'''
                UPDATE winning_wallets
                SET win_rate{time_period_suffix} = ?, 
                    balance_change{time_period_suffix} = ?, 
                    token_accounts{time_period_suffix} = ?
                WHERE wallet_address_id = ?
            '''
            self.db_cursor.execute(update_sql, (win_rate, balance_change, token_accounts, self.wallet_id))
        else:
            insert_sql = f'''
                INSERT INTO winning_wallets (
                    wallet_address_id, 
                    win_rate_7, balance_change_7, token_accounts_7, 
                    win_rate_14, balance_change_14, token_accounts_14,
                    win_rate_30, balance_change_30, token_accounts_30,
                    win_rate_60, balance_change_60, token_accounts_60,
                    win_rate_90, balance_change_90, token_accounts_90
                ) VALUES (
                    ?, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL
                )
            '''
            self.db_cursor.execute(insert_sql, (self.wallet_id,))
            update_sql = f'''
                UPDATE winning_wallets
                SET win_rate{time_period_suffix} = ?, 
                    balance_change{time_period_suffix} = ?, 
                    token_accounts{time_period_suffix} = ?
                WHERE wallet_address_id = ?
            '''
            self.db_cursor.execute(update_sql, (win_rate, balance_change, token_accounts, self.wallet_id))

        self.db_connection.commit()

    def get_summary(self, time_period):
        query = f'''
             WITH Calculations AS (
               SELECT
                 (SUM(CASE WHEN delta_sol > 0 THEN 1 ELSE 0 END) * 1.0 / COUNT(*)) * 100 AS WinRate,
                 SUM(delta_sol) AS PnL_R,
                 SUM(CASE WHEN delta_sol < 0 THEN delta_sol ELSE 0 END) AS PnL_Loss,
                 (SUM(earned_sol) / NULLIF(SUM(spent_sol), 0) - 1) * 100 AS Balance_Change,
                 COUNT(CASE WHEN scam_tokens = 1 THEN 1 END) AS ScamTokens
               FROM pnl_info
               WHERE wallet_address_id = ?
                 AND last_trade >= strftime('%s', 'now', '-{time_period} days')
             )

             SELECT
               *,
               '{time_period} days' AS TimePeriod
             FROM Calculations;
             '''
        self.db_cursor.execute(query, (self.wallet_id,))

        summary_result = self.db_cursor.fetchone()

        summary_data = {
            'SolBalance': self.sol_balance,
            'WalletAddress': str(self.wallet_address),
            'WinRate': summary_result[0],
            'PnL_R': summary_result[1],
            'PnL_Loss': summary_result[2],
            'Balance_Change': summary_result[3],
            'ScamTokens': summary_result[4],
            'TimePeriod': summary_result[5]
        }

        return summary_data

    def get_transactions(self, time_period):
        query = f'''
        SELECT *
        FROM pnl_info
        WHERE wallet_address_id = ?
          AND last_trade >= strftime('%s', 'now', '-{time_period} days');
        '''
        self.db_cursor.execute(query, (self.wallet_id,))
        results = self.db_cursor.fetchall()
        transactions_df = pd.DataFrame(results, columns=['token_account', 'wallet_address_id', 'income', 'outcome',
                                                         'total_fee', 'spent_sol', 'earned_sol', 'delta_token',
                                                         'delta_sol', 'delta_percentage', 'buys', 'sells',
                                                         'last_trade', 'time_period', 'contract', 'scam token', 'buy_period'])

        transactions_df = transactions_df.sort_values(by='last_trade', ascending=False)

        return transactions_df

    async def generate_reports_for_time_periods(self, time_periods):
        await self.initialize()

        summary = self.get_summary(time_periods[0])
        transactions = self.get_transactions(time_periods[0])
        if summary:
            file_name = f"{self.wallet_address}_{time_periods[0]}_days.xlsx"
            self.export_to_excel(summary, transactions, file_name)
            print(f"Exported summary and transactions for {time_periods[0]} days to {file_name}")
        else:
            print(f"No summary found for {time_periods[0]} days.")

    def export_to_excel(self, summary, transactions, file_name):
        sol_current_price = self.solana_usd_price
        summary['Current_Sol_Price'] = sol_current_price
        summary['ID'] = self.wallet_id
        win_rate = summary['WinRate']

        try:
            summary['Profit_USD'] = summary['PnL_R'] * sol_current_price
            summary['Loss_USD'] = abs(summary['PnL_Loss']) * sol_current_price
        except TypeError:
            summary['Profit_USD'] = 0
            summary['Loss_USD'] = 0

        summary_df = pd.DataFrame([summary], columns=['ID', 'WalletAddress', 'SolBalance', 'Current_Sol_Price', 'WinRate', 'PnL_R', 'PnL_Loss',
                                                      'Balance_Change', 'ScamTokens', 'Profit_USD',
                                                      'Loss_USD', 'TimePeriod'])

        transactions_df = pd.DataFrame(transactions, columns=['token_account', 'wallet_address_id', 'income', 'outcome',
                                                              'total_fee', 'spent_sol', 'earned_sol', 'delta_token',
                                                              'delta_sol', 'delta_percentage', 'buys', 'sells',
                                                              'last_trade', 'time_period', 'buy_period', 'contract', 'scam token'])
        transactions_df.drop(columns=['wallet_address_id'], inplace=True)
        transactions_df['last_trade'] = transactions_df['last_trade'].apply(lambda x: datetime.fromtimestamp(int(x)).strftime('%d.%m.%Y'))

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary and Transactions', index=False, startrow=0)
            row_to_start = len(summary_df) + 2
            transactions_df.to_excel(writer, sheet_name='Summary and Transactions', index=False, startrow=row_to_start)

            workbook = writer.book
            worksheet = writer.sheets['Summary and Transactions']

            for row in worksheet.iter_rows(min_row=row_to_start + 2, max_col=worksheet.max_column):
                for cell in row:
                    if cell.column == 1:
                        cell.hyperlink = f'https://solscan.io/account/{cell.value}#splTransfer'
                        cell.value = 'View Solscan'
                        cell.font = Font(underline='single')
                    elif cell.column == 15:
                        cell.hyperlink = f'https://dexscreener.com/solana/{cell.value}?maker={self.wallet_address}'
                        cell.value = 'View Dexscreener'
                        cell.font = Font(underline='single')

            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            brown_fill = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
            gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.row == 1:
                        cell.font = Font(bold=True)

            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

            for idx, row in enumerate(transactions_df.itertuples(index=False), start=row_to_start + 2):
                outcome = row[2]
                income = row[1]
                delta_percentage = row[8]
                time_period = row[12]
                buys = row[9]

                if round(outcome, 1) > round(income, 1):
                    worksheet.cell(row=idx, column=3).fill = yellow_fill

                if delta_percentage == -100:
                    worksheet.cell(row=idx, column=9).fill = brown_fill

                if pd.to_timedelta(time_period) < pd.Timedelta(minutes=1):
                    worksheet.cell(row=idx, column=13).fill = yellow_fill

                if buys > 3:
                    worksheet.cell(row=idx, column=10).fill = yellow_fill

            for idx, row in enumerate(transactions_df.itertuples(index=False), start=row_to_start + 2):
                if row[7] < 0:
                    worksheet.cell(row=idx, column=8).fill = red_fill
                elif row[7] > 0:
                    worksheet.cell(row=idx, column=8).fill = green_fill

                if row[7] < 0 and row[8] != -100:
                    worksheet.cell(row=idx, column=9).fill = red_fill
                elif row[8] > 0:
                    worksheet.cell(row=idx, column=9).fill = green_fill

            for idx, row in enumerate(summary_df.itertuples(index=False), start=1):
                if row[9] > row[10] and row[4] >= 50:
                    worksheet.cell(row=idx, column=10).fill = gold_fill
                    worksheet.cell(row=idx, column=5).fill = gold_fill

                elif row[9] < row[10] and row[4] < 50:
                    worksheet.cell(row=idx, column=10).fill = red_fill
                    worksheet.cell(row=idx, column=5).fill = red_fill

                elif row[9] < row[10] and row[4] == 50:
                    worksheet.cell(row=idx, column=10).fill = red_fill
                    worksheet.cell(row=idx, column=5).fill = gold_fill

                elif row[9] > row[10] and row[4] < 50:
                    worksheet.cell(row=idx, column=10).fill = gold_fill
                    worksheet.cell(row=idx, column=5).fill = red_fill
                else:
                    worksheet.cell(row=idx, column=11).fill = red_fill
                    worksheet.cell(row=idx, column=5).fill = red_fill

            workbook.save(file_name)

    def get_wallet_identifier(self, wallet_address):
        self.db_cursor.execute('SELECT id FROM wallet_address WHERE wallet_address = ?', (str(wallet_address),))
        result = self.db_cursor.fetchone()

        if result:
            return result[0]
        else:
            self.db_cursor.execute('INSERT INTO wallet_address (wallet_address) VALUES (?)', (str(wallet_address),))
            self.db_connection.commit()
            return self.db_cursor.lastrowid

    def convert_unix_to_date(self, unix_timestamp):
        timestamp_str = str(unix_timestamp)
        if len(timestamp_str) > 10:
            print("retruned", datetime.fromtimestamp(round(unix_timestamp / 1000)))
            return datetime.fromtimestamp(round(unix_timestamp / 1000))

        return datetime.fromtimestamp(unix_timestamp)

    async def update_token_account(self, wallet_address, wallet_token_account, block_time, wallet_address_id):
        try:
            wallet_token_account_str = str(wallet_token_account)

            self.db_cursor.execute(
                'INSERT INTO token_accounts (wallet_address_id, wallet_token_account, block_time) VALUES (?, ?, ?)',
                (wallet_address_id, wallet_token_account_str, block_time))
            self.db_connection.commit()
            print(
                f"{TerminalColors.CYAN}New token account added for wallet address: {str(wallet_address)}, token account: {wallet_token_account_str}",
                TerminalColors.RESET)

        except Exception as e:
            print(f"Error updating token account: {e}")
            self.db_connection.rollback()

    def get_token_data(self, decimals):
        for token_balances in decimals:
            if token_balances.owner == self.wallet_address and token_balances.mint != Pubkey.from_string(SOLANA_WRAPPED_MINT):
                token_contract = token_balances.mint
                token_decimal = token_balances.ui_token_amount.decimals
                if token_contract is None:
                    token_contract = "Unknown"

                return token_contract, token_decimal

    async def transactionType(self, Account: str):
        data_response = await self.solana_client.get_account_info(Pubkey.from_string(Account))
        data = data_response.value.data
        parsed_data = layouts.SPL_ACCOUNT_LAYOUT.parse(data)
        mint = Pubkey.from_bytes(parsed_data.mint)
        if mint == SOLANA_WRAPPED_MINT:
            return mint
        return mint

    async def transactionDetails(self, block_time, txn_fee, information_array: list, token_traded: str,
                           token_decimal: int):
        try:
            first_info = information_array[0]
            second_info = information_array[1]
            first_authority = first_info['authority']
            t_type = await self.transactionType(second_info['source'])

            if first_authority == str(self.wallet_address) and str(t_type) != SOLANA_WRAPPED_MINT:
                first_amount = int(first_info['amount']) / 10 ** 9
                self.update_buy(int(second_info['amount']) / 10 ** token_decimal, txn_fee, block_time)
                self.sol_spent += first_amount
                self.current_contract = token_traded
                self.last_transaction_time = block_time
                print(f"{TerminalColors.GREEN}BUY {first_amount} SOL {TerminalColors.RESET} -FOR  {int(second_info['amount']) / 10 ** token_decimal} TokenBought= {self.current_contract}  ")

            else:
                first_amount = int(first_info['amount']) / 10 ** token_decimal
                self.update_sell(first_amount, txn_fee, block_time)
                self.sol_earned += int(second_info['amount']) / 10 ** 9
                self.current_contract = token_traded
                self.last_transaction_time = block_time
                print(
                    f"{TerminalColors.RED}SELL{TerminalColors.RESET} {first_amount} -{self.current_contract}--FOR  {TerminalColors.GREEN}{int(second_info['amount']) / 10 ** 9} {TerminalColors.RESET}SOL")
        except Exception as e:
            print(f"Error processing transaction details: {e}")

    async def process_token_account(self, token_accounts: list, wallet_address_id: int):
        self.active_token_accounts = len(token_accounts)

        while self.active_token_accounts > 0:
            try:
                for token_account in token_accounts:
                    self.reset_variables()
                    print(token_account, "Number of token Account to be processed", len(token_accounts))

                    token_account_str = str(token_account)
                    sig = await self.solana_client.get_signatures_for_address(Pubkey.from_string(token_account), limit=500)
                    last_signature = await self.solana_client.get_transaction(sig.value[-1].signature,
                                                                                    encoding="jsonParsed",
                                                                                    max_supported_transaction_version=0)

                    block_time = last_signature.value.block_time
                    try:
                        await self.update_token_account(self.wallet_address, Pubkey.from_string(token_account_str), block_time, wallet_address_id)

                    except Exception as e:
                        print(f"The problem is here: {e}")

                    print("UPDATING TOKEN ACCOUNT DONE")
                    for signature in reversed(sig.value):
                        if signature.err == None:
                            transaction = await self.solana_client.get_transaction(signature.signature, encoding="jsonParsed",
                                                                                         max_supported_transaction_version=0)

                            txn_fee = transaction.value.transaction.meta.fee
                            instruction_list = transaction.value.transaction.meta.inner_instructions
                            account_signer = transaction.value.transaction.transaction.message.account_keys[0].pubkey
                            decimals = transaction.value.transaction.meta.post_token_balances
                            information_array = []

                            print(TerminalColors.RED, signature.signature, TerminalColors.RESET)
                            print(account_signer, self.wallet_address)

                            if account_signer == self.wallet_address:
                                for ui_inner_instructions in instruction_list:
                                    for txn_instructions in ui_inner_instructions.instructions:
                                        if txn_instructions.program_id == TOKEN_PROGRAM_ID:
                                            txn_information = txn_instructions.parsed['info']
                                            if 'destination' in txn_information:
                                                information_array.append(txn_information)

                                try:
                                    block_time = transaction.value.block_time

                                    token_traded, token_decimal = self.get_token_data(decimals)
                                    await self.transactionDetails(block_time, txn_fee, information_array, token_traded,
                                                                  token_decimal)
                                except Exception as e:
                                    print(e, signature.signature)
                                    print(TerminalColors.RED, "Error Adding Transaction Details", TerminalColors.RESET)
                                    continue

                    print("Calculating and updating pnl")
                    await self.calculate_deltas()
                    self.print_summary()

                    await self.fill_pnl_info_table(token_account_str, wallet_address_id)
                    token_accounts.remove(token_account)
                    self.active_token_accounts = len(token_accounts)

            except Exception as e:
                print(f"Error processing token account: {e}, {token_account}")

                if str(e) == "type NoneType doesn't define __round__ method":
                    token_accounts.remove(token_account)

                continue

    async def pair_createdTime(self, token_traded):
        url = f'https://api.dexscreener.com/latest/dex/tokens/{token_traded}'

        response = requests.get(url)
        data = response.json()
        if data['pairs'] is not None:
            return round(data['pairs'][0]['pairCreatedAt'] / 1000)
        return 0

    def calculate_time_difference(self, unix_timestamp1, unix_timestamp2):
        date1 = self.convert_unix_to_date(round(unix_timestamp1))
        date2 = self.convert_unix_to_date(round(unix_timestamp2))

        difference = date2 - date1

        hours, remainder = divmod(difference.total_seconds(), 3600)
        minutes, seconds = divmod(remainder, 60)

        if hours > 0:
            return f"{int(hours)}h {int(minutes)}m {int(seconds)}s"
        elif minutes > 0:
            return f"{int(minutes)}m {int(seconds)}s"
        else:
            return f"{int(seconds)}s"

    def update_buy(self, amount, fee, block_time):
        self.total_income += amount
        self.transaction_fees += fee
        self.buy_count += 1
        if self.initial_buy_time is None:
            self.initial_buy_time = block_time

    def update_sell(self, amount, fee, block_time):
        self.total_outcome += amount
        self.transaction_fees += fee
        self.sell_count += 1
        self.final_sell_time = block_time
        if self.last_transaction_time is None or block_time > self.last_transaction_time:
            self.last_transaction_time = block_time

    def print_summary(self):
        print(f"Income: {self.total_income}")
        print(f"Outcome: {self.total_outcome}")
        print(f"Total Fee: {self.transaction_fees / SOLANA_DECIMALS}")
        print(f"Spent SOL: {self.sol_spent}")
        print(f"Earned SOL: {self.sol_earned}")
        print(f"Delta Token: {self.token_difference}")
        print(f"Delta SOL: {self.sol_difference}")
        print(f"Delta Percentage: {self.profit_percentage}%")
        print(f"Buys: {self.buy_count}")
        print(f"Sells: {self.sell_count}")

        print(f"Time Period: {self.trading_duration}")
        print(f"Contract: {self.current_contract}")
        print(f"Scam Tokens: {self.suspicious_tokens}")

        if self.token_creation_time == 0:
            buy_period = "Unknown"
            print("Buy Period:", buy_period)
            print(self.current_contract)

        else:
            try:
                buy_period = self.calculate_time_difference(self.token_creation_time, self.initial_buy_time)
                print("Buy Period:", buy_period)
            except Exception as e:
                print(f"Error calculating buy period: {e}")
                buy_period = "No buy"

    def reset_variables(self):
        self.total_income = 0
        self.total_outcome = 0
        self.transaction_fees = 0
        self.sol_spent = 0
        self.sol_earned = 0
        self.buy_count = 0
        self.sell_count = 0
        self.sol_difference = 0
        self.token_difference = 0
        self.profit_percentage = 0
        self.initial_buy_time = None
        self.final_sell_time = None
        self.last_transaction_time = None
        self.trading_duration = 0
        self.current_contract = None
        self.suspicious_tokens = 0
        self.token_creation_time = 0
        self.purchase_period = 0

    async def getToken_SolAmount(self):
        token_address = str(self.current_contract)
        url = f'https://api.dexscreener.com/latest/dex/tokens/{token_address}'

        response = requests.get(url)
        data = response.json()
        if data['pairs'] is not None:
            token_price_usd = float(data['pairs'][0]['priceUsd'])
            wallet_amount = self.total_income
            Worth_wallet_amount = token_price_usd * wallet_amount
            worth_in_solana = Worth_wallet_amount / self.solana_usd_price
            return worth_in_solana
        else:
            self.suspicious_tokens = 1
            return 0

    async def calculate_deltas(self):
        if self.sell_count > 0:
            self.token_difference = self.total_income - self.total_outcome
            self.sol_difference = self.sol_earned - self.sol_spent
            self.profit_percentage = (self.sol_difference / self.sol_spent) * 100 if self.sol_spent != 0 else 0
        else:
            self.token_difference = self.total_income
            self.sol_difference = self.sol_earned - self.sol_spent
            self.profit_percentage = -100

        self.token_creation_time = await self.pair_createdTime(self.current_contract)
        if self.token_creation_time == 0:
            self.purchase_period = "Unknown"

        else:
            self.purchase_period = self.calculate_time_difference(self.token_creation_time, self.initial_buy_time)

        if self.initial_buy_time and self.final_sell_time:
            time_difference = self.final_sell_time - self.initial_buy_time
            self.trading_duration = self.calculate_time_difference(self.initial_buy_time, self.final_sell_time)
        elif self.initial_buy_time and not self.final_sell_time:
            self.trading_duration = 0
            self.suspicious_tokens = 1

    async def fill_pnl_info_table(self, token_account, wallet_address_id):
        try:
            fields = [
                ('token_account', token_account),
                ('income', self.total_income),
                ('outcome', self.total_outcome),
                ('fee', self.transaction_fees),
                ('spent_sol', self.sol_spent),
                ('earned_sol', self.sol_earned),
                ('delta_token', self.token_difference),
                ('delta_sol', self.sol_difference),
                ('buys', self.buy_count),
                ('sells', self.sell_count),
                ('time_period', self.trading_duration),
                ('contract', self.current_contract),
                ('scam_tokens', self.suspicious_tokens),
                ('wallet_address_id', wallet_address_id),
                ('last_trade', self.last_transaction_time),
                ('buy_period', self.purchase_period)
            ]

            none_fields = [field_name for field_name, field_value in fields if field_value is None]
            if none_fields:
                print(f"One or more fields are None: {', '.join(none_fields)}. Skipping the operation.")
                return

            insert_sql = '''
                    INSERT INTO pnl_info (
                        wallet_address_id, token_account, income, outcome, total_fee, spent_sol, earned_sol,
                        delta_token, delta_sol, delta_percentage, buys, sells, last_trade,
                        time_period, contract, scam_tokens, buy_period
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?)
                '''

            if any(value is None for value in
                   [token_account, self.total_income, self.total_outcome, self.transaction_fees, self.sol_spent, self.sol_earned, self.token_difference,
                    self.sol_difference, self.buy_count, self.sell_count, self.trading_duration, self.current_contract, self.suspicious_tokens,
                    wallet_address_id, self.last_transaction_time, self.calculate_time_difference(self.initial_buy_time, self.token_creation_time) if self.token_creation_time != 0 or self.initial_buy_time != None else 0]):
                print("One or more fields are None. Skipping the operation.")
                return

            check_sql = '''
                    SELECT   1 FROM pnl_info
                    WHERE wallet_address_id = ? AND last_trade = ?
                '''
            cursor = self.db_connection.cursor()
            cursor.execute(check_sql, (wallet_address_id, self.last_transaction_time))
            if cursor.fetchone() is not None:
                print("Transaction already exists in the database. Skipping the operation.")
                return

            insert_data = (
                wallet_address_id,
                token_account,
                self.total_income,
                self.total_outcome,
                self.transaction_fees / SOLANA_DECIMALS,
                self.sol_spent,
                self.sol_earned,
                self.token_difference,
                self.sol_difference,
                self.profit_percentage,
                self.buy_count,
                self.sell_count,
                self.last_transaction_time,
                self.trading_duration,
                str(self.current_contract),
                self.suspicious_tokens,
                self.purchase_period
            )

            cursor.execute(insert_sql, insert_data)
            self.db_connection.commit()

            print("PNL info successfully inserted into the database.")
        except Exception as e:
            print(f"Filling info issue, {e}")

    async def process_transactions(self):
        try:
            wallet_address_id = self.wallet_id

            owner = self.wallet_address
            opts = TokenAccountOpts(
                program_id=Pubkey.from_string("TokenkegQfeZyiNwAJbNbGKPFXCWuBvf9Ss623VQ5DA")
            )
            response = await self.solana_client.get_token_accounts_by_owner(owner, opts)
            solana_token_accounts = {str(token_account.pubkey): token_account for token_account in response.value}

            num_tokenAccounts = len(solana_token_accounts)
            print("Number of token Accounts", num_tokenAccounts)
            self.db_cursor.execute('SELECT wallet_token_account FROM token_accounts WHERE wallet_address_id = ?',
                           (wallet_address_id,))
            db_token_accounts = {row[0] for row in self.db_cursor.fetchall()}
            newTokenAccounts = solana_token_accounts.keys() - db_token_accounts
            new_token_accounts = list(newTokenAccounts)

            if not newTokenAccounts:
                print("No new token accounts")
                return

            if 1 <= len(newTokenAccounts) < MAX_TOKEN_ACCOUNTS:
                print(
                    f"Processing Address {self.wallet_address} Number of Token Accounts to be Processed {len(newTokenAccounts)}")
                new_token_accounts = list(newTokenAccounts)
                await self.process_token_account(new_token_accounts, wallet_address_id)
                print("ALL TOKEN ACCOUNTS PROCESSED")
            else:
                print(
                    f"{TerminalColors.RED}Skipping Wallet Address {self.wallet_address} with {len(newTokenAccounts)} token accounts",
                    TerminalColors.RESET)

        except Exception as e:
            print(e)
            print(f"Failed to process transaction {self.wallet_address} {e}")

async def run():
    processor = SolanaTrader(Pubkey.from_string("EWzk2847WCPis45hPQ6Em5UuRbZiP1CSmqx7nG9ELd2L"))

    await processor.initialize()

    await processor.process_transactions()

    await processor.generate_reports_for_time_periods([90, 60, 30, 14, 7, 1])

    print("All accounts processed and reports generated.")

asyncio.run(run())